# -*- coding: utf-8 -*-
"""Gestión de asistencia HMX: captura de supervisores + cruce con reloj checador.

Flujo:
1. Los supervisores capturan la asistencia diaria por empleado (tipo de
   incidencia, turno, máquina y tiempo extra). Cada movimiento queda en el
   chatter con hora y autor.
2. RH sube el archivo que genera el reloj checador (asistencia.xls/.xlsx,
   columnas: Número, Nombre, Tiempo, Estado, Dispositivos, Tipo de Registro).
   Las checadas se empatan contra el número de nómina del empleado.
3. El cruce toma la primera checada del día como entrada y la última como
   salida (la columna "Estado" del dispositivo no es confiable) y clasifica
   cada registro: coincide, sin checada, checó con incidencia capturada,
   checó sin captura (se genera el registro para validar) o checada única.
"""
import base64
import io
import logging
import re
from collections import defaultdict
from datetime import datetime

import pytz

from odoo import api, fields, models, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

# Las checadas del archivo vienen en hora local de las plantas. Desde 2022
# todas las sedes (Monterrey, Querétaro, Tlaxcala) comparten UTC-6 fijo.
CLOCK_TZ = 'America/Mexico_City'

CROSS_STATUS = [
    ('pendiente', 'Sin cruzar'),
    ('ok', 'Coincide'),
    ('sin_checada', 'Capturado presente, sin checada'),
    ('checo_con_incidencia', 'Checó pero se capturó incidencia'),
    ('sin_captura', 'Checó sin captura del supervisor'),
    ('incompleta', 'Checada única (entrada o salida faltante)'),
]


def _clock_to_utc(dt):
    """Interpreta un datetime naive como hora local de planta y regresa naive UTC."""
    return pytz.timezone(CLOCK_TZ).localize(dt).astimezone(pytz.utc).replace(tzinfo=None)


def _utc_to_clock(dt):
    """Convierte un datetime naive UTC (almacenado) a hora local de planta."""
    return pytz.utc.localize(dt).astimezone(pytz.timezone(CLOCK_TZ)).replace(tzinfo=None)


class HmxAttendanceIncidenceType(models.Model):
    _name = 'hmx.attendance.incidence.type'
    _description = 'Tipo de incidencia de asistencia HMX'
    _order = 'sequence, id'

    name = fields.Char('Nombre', required=True, translate=False)
    code = fields.Char('Código', required=True, help='Código usado en las listas de asistencia (A, F, V, ...).')
    sequence = fields.Integer(default=10)
    active = fields.Boolean(default=True)
    is_attendance = fields.Boolean(
        'Cuenta como asistencia',
        help='Marcado para tipos donde se espera que el empleado haya checado (A).'
    )
    justifies_absence = fields.Boolean(
        'Justifica la ausencia',
        help='Marcado para permisos, vacaciones, incapacidades y suspensiones: '
             'no se espera checada y no se considera discrepancia.'
    )
    color = fields.Integer('Color')

    _sql_constraints = [
        ('code_uniq', 'unique(code)', 'Ya existe un tipo de incidencia con ese código.'),
    ]

    def name_get(self):
        return [(rec.id, f"{rec.code} - {rec.name}") for rec in self]


class HmxAttendanceRecord(models.Model):
    _name = 'hmx.attendance.record'
    _description = 'Registro diario de asistencia HMX'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'date desc, employee_id'
    _rec_name = 'display_name'

    display_name = fields.Char(compute='_compute_display_name', store=True)
    employee_id = fields.Many2one(
        'hr.employee', 'Empleado', required=True, index=True, tracking=True,
        ondelete='restrict',
    )
    numero_nomina = fields.Integer(
        related='employee_id.x_numero_nomina', string='Nómina', store=True,
    )
    planta = fields.Selection(
        related='employee_id.x_planta', string='Planta', store=True,
    )
    date = fields.Date('Fecha', required=True, index=True, tracking=True,
                       default=fields.Date.context_today)
    incidence_type_id = fields.Many2one(
        'hmx.attendance.incidence.type', 'Incidencia', required=True, tracking=True,
        default=lambda self: self.env.ref(
            'empleados_hmx.incidence_type_asistencia', raise_if_not_found=False),
    )
    turno = fields.Selection(
        [('dia', 'Día'), ('tarde', 'Tarde'), ('noche', 'Noche'), ('mixto', 'Mixto')],
        string='Turno', tracking=True,
    )
    maquina = fields.Char('Máquina / Área', tracking=True)
    overtime_hours = fields.Float('Tiempo extra (hrs)', tracking=True)
    notes = fields.Char('Observaciones', tracking=True)

    # Rastro de la captura: quién y exactamente cuándo hizo el movimiento.
    captured_by_id = fields.Many2one(
        'res.users', 'Capturado por', readonly=True,
        default=lambda self: self.env.user,
    )
    captured_at = fields.Datetime(
        'Hora de captura', readonly=True, default=fields.Datetime.now,
    )
    source = fields.Selection(
        [('supervisor', 'Captura de supervisor'), ('checador', 'Generado por cruce del checador')],
        string='Origen', default='supervisor', readonly=True, tracking=True,
    )
    session_id = fields.Many2one(
        'hmx.attendance.session', 'Sesión de captura', readonly=True,
        ondelete='set null',
    )

    # Resultado del cruce con el reloj checador.
    checador_entry = fields.Datetime('Entrada (checador)', readonly=True)
    checador_exit = fields.Datetime('Salida (checador)', readonly=True)
    checador_hours = fields.Float(
        'Horas checadas', compute='_compute_checador_hours', store=True,
    )
    cross_status = fields.Selection(
        CROSS_STATUS, string='Resultado del cruce', default='pendiente',
        readonly=True, tracking=True, index=True,
    )
    clock_import_id = fields.Many2one(
        'hmx.attendance.clock.import', 'Importación del cruce', readonly=True,
    )
    state = fields.Selection(
        [('draft', 'Capturado'), ('to_review', 'Por validar'), ('validated', 'Validado')],
        string='Estado', default='draft', tracking=True, index=True,
    )

    _sql_constraints = [
        ('employee_date_uniq', 'unique(employee_id, date)',
         'Ya existe un registro de asistencia para ese empleado en esa fecha.'),
    ]

    @api.depends('employee_id', 'date')
    def _compute_display_name(self):
        for rec in self:
            rec.display_name = (
                f"{rec.employee_id.name or ''} - {rec.date or ''}" if rec.employee_id else _('Nuevo')
            )

    @api.depends('checador_entry', 'checador_exit')
    def _compute_checador_hours(self):
        for rec in self:
            if rec.checador_entry and rec.checador_exit and rec.checador_exit > rec.checador_entry:
                rec.checador_hours = (rec.checador_exit - rec.checador_entry).total_seconds() / 3600.0
            else:
                rec.checador_hours = 0.0

    @api.constrains('date')
    def _check_date(self):
        for rec in self:
            if rec.date and rec.date > fields.Date.context_today(rec):
                raise ValidationError(_('No se puede capturar asistencia de una fecha futura.'))

    def write(self, vals):
        # Toda edición de campos de captura re-sella quién y cuándo hizo el movimiento.
        capture_fields = {'incidence_type_id', 'turno', 'maquina', 'overtime_hours', 'notes', 'date', 'employee_id'}
        if capture_fields & set(vals):
            validated = self.filtered(lambda r: r.state == 'validated')
            if validated and not self.env.user.has_group('empleados_hmx.group_hmx_attendance_manager'):
                raise UserError(_(
                    'Un registro validado solo lo puede modificar el gerente de asistencia '
                    '(registros: %s).'
                ) % ', '.join(validated.mapped('display_name')))
            vals = dict(vals, captured_by_id=self.env.user.id, captured_at=fields.Datetime.now())
        return super().write(vals)

    def action_validate(self):
        if not self.env.user.has_group('empleados_hmx.group_hmx_attendance_manager'):
            raise UserError(_('Solo el gerente de asistencia puede validar registros.'))
        self.write({'state': 'validated'})

    def action_reset_draft(self):
        self.write({'state': 'draft'})


class HmxAttendanceClockLine(models.Model):
    _name = 'hmx.attendance.clock.line'
    _description = 'Checada importada del reloj checador'
    _order = 'punch_time'

    import_id = fields.Many2one(
        'hmx.attendance.clock.import', 'Importación', required=True,
        index=True, ondelete='cascade',
    )
    employee_number = fields.Integer('Número (nómina)', required=True, index=True)
    employee_name_file = fields.Char('Nombre en archivo')
    employee_id = fields.Many2one('hr.employee', 'Empleado', index=True)
    punch_time = fields.Datetime('Checada', required=True)
    punch_date = fields.Date('Fecha (local)', compute='_compute_punch_date', store=True)
    device = fields.Char('Dispositivo')
    raw_state = fields.Char('Estado (archivo)', help='Etiqueta del dispositivo; no es confiable para entrada/salida.')

    _sql_constraints = [
        ('punch_uniq', 'unique(employee_number, punch_time, device)',
         'Checada duplicada: ese empleado ya tiene registrado ese momento en ese dispositivo.'),
    ]

    @api.depends('punch_time')
    def _compute_punch_date(self):
        for line in self:
            line.punch_date = _utc_to_clock(line.punch_time).date() if line.punch_time else False


class HmxAttendanceClockImport(models.Model):
    _name = 'hmx.attendance.clock.import'
    _description = 'Importación de archivo del reloj checador'
    _inherit = ['mail.thread']
    _order = 'create_date desc'

    name = fields.Char('Referencia', default=lambda self: _('Nueva importación'), readonly=True, copy=False)
    file_data = fields.Binary('Archivo del checador', attachment=True)
    file_name = fields.Char('Nombre del archivo')
    state = fields.Selection(
        [('draft', 'Borrador'), ('imported', 'Checadas importadas'), ('crossed', 'Cruce generado')],
        default='draft', tracking=True,
    )
    line_ids = fields.One2many('hmx.attendance.clock.line', 'import_id', 'Checadas')
    line_count = fields.Integer(compute='_compute_counts')
    matched_count = fields.Integer('Checadas empatadas', compute='_compute_counts')
    date_from = fields.Date('Desde', readonly=True)
    date_to = fields.Date('Hasta', readonly=True)
    skipped_duplicates = fields.Integer('Duplicadas omitidas', readonly=True)
    unmatched_numbers = fields.Text(
        'Nóminas sin empleado', readonly=True,
        help='Números del archivo que no corresponden a ningún empleado con ese número de nómina.',
    )
    record_ids = fields.One2many('hmx.attendance.record', 'clock_import_id', 'Registros cruzados')
    record_count = fields.Integer(compute='_compute_counts')

    @api.depends('line_ids', 'line_ids.employee_id', 'record_ids')
    def _compute_counts(self):
        for imp in self:
            imp.line_count = len(imp.line_ids)
            imp.matched_count = len(imp.line_ids.filtered('employee_id'))
            imp.record_count = len(imp.record_ids)

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if not vals.get('name') or vals['name'] == _('Nueva importación'):
                vals['name'] = self.env['ir.sequence'].next_by_code('hmx.attendance.clock.import') or '/'
        return super().create(vals_list)

    # ------------------------------------------------------------------
    # Lectura del archivo
    # ------------------------------------------------------------------
    def _read_rows(self):
        """Regresa las filas del archivo como listas de valores."""
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('Adjunta el archivo generado por el reloj checador.'))
        content = base64.b64decode(self.file_data)
        fname = (self.file_name or '').lower()
        if fname.endswith('.xlsx'):
            try:
                import openpyxl
            except ImportError:
                raise UserError(_('El servidor no tiene instalada la librería openpyxl.'))
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = wb[wb.sheetnames[0]]
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        if fname.endswith('.xls'):
            try:
                import xlrd
            except ImportError:
                raise UserError(_('El servidor no tiene instalada la librería xlrd para leer .xls.'))
            wb = xlrd.open_workbook(file_contents=content)
            sheet = wb.sheet_by_index(0)
            rows = []
            for r in range(sheet.nrows):
                row = []
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        row.append(xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode))
                    else:
                        row.append(cell.value)
                rows.append(row)
            return rows
        raise UserError(_('Formato no soportado: sube el .xls o .xlsx que exporta el checador.'))

    @staticmethod
    def _parse_punch_time(value):
        """El checador exporta 'dd/mm/aaaa hh:mm:ss a. m./p. m.' (12 horas, locale
        mexicano); otros exports traen 24 horas o datetime directo."""
        if isinstance(value, datetime):
            return value
        text = str(value or '').strip()
        # Normaliza 'a. m.', 'a.m.', 'A. M.' → 'AM' para poder usar %p.
        normalized = re.sub(r'\s*([aApP])\.?\s*[mM]\.?\s*$', lambda m: ' ' + m.group(1).upper() + 'M', text)
        for fmt in ('%d/%m/%Y %I:%M:%S %p', '%d/%m/%Y %I:%M %p',
                    '%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(normalized, fmt)
            except ValueError:
                continue
        return None

    @api.model
    def _employee_map(self):
        """Mapa número de nómina → empleado, la conexión Odoo ↔ checador.

        Incluye archivados (sus checadas históricas deben empatar), pero si un
        número existiera repetido entre archivado y activo, prevalece el activo.
        La unicidad del número se garantiza con la restricción en hr.employee.
        """
        employees = self.env['hr.employee'].with_context(active_test=False).search([
            ('x_numero_nomina', '!=', 0),
        ], order='active asc')  # los activos van al final y prevalecen en el mapa
        return {emp.x_numero_nomina: emp for emp in employees}

    def action_rematch_lines(self):
        """Reintenta la conexión de checadas sin empleado.

        Flujo: el archivo trae números de nómina que no existen en Odoo → se
        capturan en la ficha del empleado (Número de nómina) → este botón
        vuelve a empatar sin necesidad de reimportar el archivo.
        """
        self.ensure_one()
        pending = self.line_ids.filtered(lambda l: not l.employee_id)
        if not pending:
            raise UserError(_('Todas las checadas de esta importación ya tienen empleado.'))
        emp_by_number = self._employee_map()
        matched = 0
        for line in pending:
            employee = emp_by_number.get(line.employee_number)
            if employee:
                line.employee_id = employee.id
                matched += 1
        still = sorted({l.employee_number for l in self.line_ids if not l.employee_id})
        self.write({'unmatched_numbers': ', '.join(str(n) for n in still) or False})
        self.message_post(body=_(
            'Reintento de empate: %(matched)s checadas conectadas; '
            '%(missing)s números siguen sin empleado.'
        ) % {'matched': matched, 'missing': len(still)})
        return True

    def action_import_file(self):
        self.ensure_one()
        rows = self._read_rows()

        # Localizar el encabezado (Número / Tiempo) sin asumir la fila exacta.
        header_idx = None
        for i, row in enumerate(rows[:10]):
            cells = [str(v or '').strip().lower() for v in row]
            if any(c in ('número', 'numero') for c in cells) and 'tiempo' in cells:
                header_idx = i
                break
        if header_idx is None:
            raise UserError(_(
                'No se encontró el encabezado esperado (Número, Nombre, Tiempo, ...). '
                'Verifica que sea el archivo exportado por el reloj checador.'
            ))
        header = [str(v or '').strip().lower() for v in rows[header_idx]]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        col_num = col('número', 'numero')
        col_name = col('nombre')
        col_time = col('tiempo')
        col_state = col('estado')
        col_device = col('dispositivos', 'dispositivo')

        emp_by_number = self._employee_map()

        Line = self.env['hmx.attendance.clock.line']
        existing = {
            (l.employee_number, l.punch_time, l.device or '')
            for l in Line.search([])
        }

        to_create, unmatched, skipped, bad_rows = [], set(), 0, 0
        for row in rows[header_idx + 1:]:
            if not row or col_num >= len(row):
                continue
            raw_num = row[col_num]
            try:
                number = int(float(str(raw_num).strip()))
            except (TypeError, ValueError):
                continue
            punch_local = self._parse_punch_time(row[col_time] if col_time is not None else None)
            if not punch_local:
                bad_rows += 1
                continue
            punch_utc = _clock_to_utc(punch_local)
            device = str(row[col_device] or '').strip() if col_device is not None else ''
            key = (number, punch_utc, device)
            if key in existing:
                skipped += 1
                continue
            existing.add(key)
            employee = emp_by_number.get(number)
            if not employee:
                unmatched.add(number)
            to_create.append({
                'import_id': self.id,
                'employee_number': number,
                'employee_name_file': str(row[col_name] or '').strip() if col_name is not None else '',
                'employee_id': employee.id if employee else False,
                'punch_time': punch_utc,
                'device': device,
                'raw_state': str(row[col_state] or '').strip() if col_state is not None else '',
            })

        if not to_create and not skipped:
            raise UserError(_('El archivo no contiene checadas legibles.'))

        lines = Line.create(to_create)
        dates = lines.mapped('punch_date')
        self.write({
            'state': 'imported',
            'skipped_duplicates': skipped,
            'date_from': min(dates) if dates else False,
            'date_to': max(dates) if dates else False,
            'unmatched_numbers': ', '.join(str(n) for n in sorted(unmatched)) or False,
        })
        self.message_post(body=_(
            'Importación: %(new)s checadas nuevas, %(dup)s duplicadas omitidas, '
            '%(bad)s filas ilegibles, %(unmatched)s números de nómina sin empleado.'
        ) % {'new': len(lines), 'dup': skipped, 'bad': bad_rows, 'unmatched': len(unmatched)})
        return True

    # ------------------------------------------------------------------
    # Cruce contra la captura de los supervisores
    # ------------------------------------------------------------------
    def action_cross_check(self):
        self.ensure_one()
        if self.state == 'draft':
            raise UserError(_('Primero importa el archivo del checador.'))
        if not self.line_ids.filtered('employee_id'):
            raise UserError(_('Ninguna checada empató con un empleado; revisa los números de nómina.'))

        # Se cruzan TODAS las checadas conectadas del rango de fechas, sin
        # importar en qué importación llegaron: si las plantas suben archivos
        # separados de la misma semana, el cruce las considera juntas.
        lines = self.env['hmx.attendance.clock.line'].search([
            ('employee_id', '!=', False),
            ('punch_date', '>=', self.date_from),
            ('punch_date', '<=', self.date_to),
        ])

        punches = defaultdict(list)
        for line in lines:
            punches[(line.employee_id.id, line.punch_date)].append(line.punch_time)

        Record = self.env['hmx.attendance.record']
        touched = Record.browse()
        created = 0
        for (emp_id, day), times in punches.items():
            times.sort()
            entry = times[0]
            exit_ = times[-1] if len(times) > 1 else False
            vals = {
                'checador_entry': entry,
                'checador_exit': exit_,
                'clock_import_id': self.id,
            }
            record = Record.search([
                ('employee_id', '=', emp_id), ('date', '=', day),
            ], limit=1)
            if record:
                # La discrepancia de fondo pesa más que la checada incompleta:
                # si se capturó incidencia (falta, permiso...) y aun así checó,
                # eso es lo que hay que revisar.
                if not record.incidence_type_id.is_attendance:
                    status = 'checo_con_incidencia'
                elif not exit_:
                    status = 'incompleta'
                else:
                    status = 'ok'
                record.write(dict(vals, cross_status=status))
            else:
                # Checó pero nadie lo capturó: se genera el registro para validar.
                record = Record.create(dict(
                    vals,
                    employee_id=emp_id,
                    date=day,
                    source='checador',
                    state='to_review',
                    cross_status='incompleta' if not exit_ else 'sin_captura',
                ))
                created += 1
            touched |= record

        # Registros capturados en el rango que no tuvieron ninguna checada.
        pending = Record.search([
            ('date', '>=', self.date_from), ('date', '<=', self.date_to),
            ('id', 'not in', touched.ids),
        ])
        no_punch = Record.browse()
        for record in pending:
            if not record.employee_id.x_numero_nomina:
                # Sin número de nómina no hay conexión con el checador:
                # no se puede conciliar ni acusar "sin checada".
                continue
            if record.incidence_type_id.is_attendance:
                record.write({'cross_status': 'sin_checada', 'clock_import_id': self.id})
                no_punch |= record
            else:
                # Falta, permiso, vacaciones, etc. sin checada: la captura
                # es congruente con el checador.
                record.write({'cross_status': 'ok', 'clock_import_id': self.id})

        self.write({'state': 'crossed'})
        self.message_post(body=_(
            'Cruce generado: %(total)s registros conciliados, %(created)s generados '
            'automáticamente por checada sin captura, %(nopunch)s capturados presentes sin checada.'
        ) % {'total': len(touched), 'created': created, 'nopunch': len(no_punch)})
        return self.action_view_records()

    def action_view_records(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Registros del cruce'),
            'res_model': 'hmx.attendance.record',
            'view_mode': 'list,form',
            'domain': [('clock_import_id', '=', self.id)],
            'context': {'search_default_group_cross_status': 1},
        }
